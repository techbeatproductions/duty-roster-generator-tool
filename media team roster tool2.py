import pandas as pd
import calendar
from datetime import datetime
import random
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

def get_sundays(year, month):
    """Returns a list of all Sundays in the given month, formatted with ordinals."""
    c = calendar.Calendar(firstweekday=calendar.SUNDAY)
    month_dates = c.itermonthdates(year, month)
    sundays = []
    for date in month_dates:
        if date.weekday() == 6 and date.month == month:
            # Get the day and month name
            day = date.day
            month_name = calendar.month_name[month]
            # Determine the ordinal suffix
            if 10 <= day % 100 <= 20:
                suffix = 'th'
            else:
                suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')
            formatted_date = f"{day}{suffix} {month_name}"
            sundays.append(formatted_date)
    return sundays

def assign_members(sundays, team_members):
    """Assigns team members to Camera, Projection, and Stream stations for each service."""
    roster = {}

    for formatted_sunday in sundays:
        # Split the formatted date into day and month name for storage
        day, month_name = formatted_sunday.split(' ', 1)
        sunday_key = f"{day} {month_name}"  # Key to be used in the roster
        roster[sunday_key] = {
            '1st Service': {'Camera': [], 'Projection': '', 'Stream': ''},
            '2nd Service': {'Camera': [], 'Projection': '', 'Stream': ''},
        }

        # Assign members for the 1st service
        members_cycle_1st = team_members.copy()
        random.shuffle(members_cycle_1st)

        # Ensure we have enough members for 1st service
        if len(members_cycle_1st) >= 4:
            roster[sunday_key]['1st Service']['Camera'] = members_cycle_1st[:2]  # Two for Camera
            roster[sunday_key]['1st Service']['Projection'] = members_cycle_1st[2]  # One for Projection
            roster[sunday_key]['1st Service']['Stream'] = members_cycle_1st[3]  # One for Stream

            # Prepare for the 2nd service, removing those already assigned
            members_cycle_2nd = members_cycle_1st[4:]  # Start from the 5th member
            random.shuffle(members_cycle_2nd)

            # Ensure we have enough members for 2nd service
            if len(members_cycle_2nd) >= 4:
                roster[sunday_key]['2nd Service']['Camera'] = members_cycle_2nd[:2]  # Two for Camera
                roster[sunday_key]['2nd Service']['Projection'] = members_cycle_2nd[2]  # One for Projection
                roster[sunday_key]['2nd Service']['Stream'] = members_cycle_2nd[3]  # One for Stream

    return roster

def create_excel_roster(year, roster, team_members, file_name='Media_Team_Roster.xlsx'):
    """Creates or updates an Excel file with a new sheet for the month, with merged date cells, wrapped text, bold headers, and adjusted column width."""
    output_path = os.path.abspath(file_name)

    try:
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            for month in range(1, 13):
                sundays = get_sundays(year, month)
                if not sundays:
                    print(f"No Sundays found in {calendar.month_name[month]} {year}.")
                    continue  # Skip this month if there are no Sundays

                data = []

                # Prepare the first row with formatted Sunday dates
                header_row_1 = ['MEMBERS', '']  # Start with 'MEMBERS' and 'STATION' columns
                for formatted_sunday in sundays:
                    header_row_1.append('')
                    header_row_1.append(formatted_sunday.upper())  # Use the formatted date
                    header_row_1.append('')  # Blank for the 2nd service cell
                    header_row_1.append('')  # Add a blank column for spacing

                data.append(header_row_1)

                # Prepare the second row with service labels (1st and 2nd Service)
                header_row_2 = ['', '']  # Start with empty cells for alignment
                for _ in sundays:
                    header_row_2.append('Media Role')
                    header_row_2.append('1ST SERVICE')
                    header_row_2.append('2ND SERVICE')
                    header_row_2.append('')  # Empty column for spacing

                data.append([item.upper() for item in header_row_2])  # Capitalize service labels

                # Add Camera, Projection, and Stream assignments (capitalize stations)
                for station in ['Camera', 'Projection', 'Stream']:
                    # Prepare rows for each station with the station name first
                    row = ['', '']  # Row label for station (capitalized)

                    for formatted_sunday in sundays:
                        sunday_key = formatted_sunday  # Use the formatted date directly

                        # Assignments for 1st Service
                        assignment_1st = ', '.join(roster[sunday_key]['1st Service'][station]).upper() if isinstance(roster[sunday_key]['1st Service'][station], list) else roster[sunday_key]['1st Service'][station].upper()
                        row.append(station.upper()) #Adding station name
                        row.append(assignment_1st) #Adding 1st Service assignment

                        # Assignments for 2nd Service
                        assignment_2nd = ', '.join(roster[sunday_key]['2nd Service'][station]).upper() if isinstance(roster[sunday_key]['2nd Service'][station], list) else roster[sunday_key]['2nd Service'][station].upper()
                        row.append(assignment_2nd)

                        

                        row.append('')  # Empty column for spacing

                    data.append(row)  # Append the row for the station after processing all Sundays


                # Add an empty row before the notes
                data.append([''] * (len(sundays) * 3 + 2))  # Adjusted for extra 'STATION' column

                # Add notes after all service assignments (capitalized)
                note_row_camera = ['']*3 +['CAMERA - THE CAMERA SHOULD NOT BE LEFT UNATTENDED AT ANY POINT DURING THE SERVICE.'] + [''] * (len(sundays) * 3 )
                note_row_projection =['']*3 + ['PROJECTION: ALL SONG LYRICS SHOULD BE DISPLAYED ON THE SCREEN.'] + [''] * (len(sundays) * 3 )
                note_row_stream = ['']*3+['STREAM: MONITORING THE STREAM HEALTH THROUGHOUT THE SERVICE, ALSO ADDING BIBLE VERSES AND SONG LYRICS TO THE LIVESTREAM.'] + [''] * (len(sundays) * 3 )

                data.append(note_row_camera)
                data.append([''] * (len(sundays) * 3 + 2))  # Empty row after Camera note
                data.append(note_row_projection)
                data.append([''] * (len(sundays) * 3 + 2))  # Empty row after Projection note
                data.append(note_row_stream)
                data.append([''] * (len(sundays) * 3 + 2))  # Empty row after Stream note

                # Create DataFrame and save to Excel
                df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name=calendar.month_name[month], index=False, header=False)

        # Open the workbook and access the worksheet to apply text wrapping, bold, and merge cells
        workbook = openpyxl.load_workbook(output_path)
        for month in range(1, 13):
            sheet = workbook[calendar.month_name[month]]

            sheet.column_dimensions['D'].width = 30

            # Merge the date cells to span the 1st and 2nd Service
            for col_idx in range(3, len(sundays) * 3 + 3, 3):  # Start at 3 to skip 'MEMBERS' and 'STATION' columns
                col_letter = get_column_letter(col_idx)  # Calculate the column letter for each date
                sheet.merge_cells(f'{col_letter}1:{col_letter}1')  # Only merge the date header
                sheet[f'{col_letter}1'].alignment = Alignment(horizontal='center', vertical='center')  # Center alignment

            # Add team members to the first column and capitalize them
            for idx, member in enumerate(team_members):
                sheet[f'A{idx + 3}'] = member.upper()  # Starting from row 3 to leave space for headers, capitalize members

            # Apply bold formatting to headers (1st and 2nd row)
            for cell in sheet[1]:  # First header row
                cell.font = Font(bold=True)
            for cell in sheet[2]:  # Second header row
                cell.font = Font(bold=True)

            # Apply text wrapping for all cells
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)

            # Conditional formatting for "1ST SERVICE" and "2ND SERVICE"
            for col in range(3, len(sundays) * 4 + 3):  # Adjust according to your data
                cell_media_role_camera = sheet.cell (row=3, column=col - 1)
                cell_media_role_projection = sheet.cell (row=4, column=col - 1)
                cell_media_role_stream = sheet.cell (row=5, column=col - 1)


                cell_1st_service = sheet.cell(row=2, column=col)  # Row 2 for "1ST SERVICE"
                cell_2nd_service = sheet.cell(row=2, column=col + 1)  # Row 2 for "2ND SERVICE"

                if cell_1st_service.value == "1ST SERVICE":
                    cell_1st_service.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Black fill
                    cell_1st_service.font = Font(color="FFFFFF")  # White font color

                if cell_2nd_service.value == "2ND SERVICE":
                    cell_2nd_service.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Black fill
                    cell_2nd_service.font = Font(color="FFFFFF")  # White font color

                if cell_media_role_camera.value == "CAMERA":
                    cell_media_role_camera.fill = PatternFill(start_color= "9933ff", end_color= "9933ff", fill_type="solid")   
                    cell_media_role_camera.font = Font(color="FFFFFF")  # White font color

                if cell_media_role_projection.value == "PROJECTION":
                    cell_media_role_projection.fill = PatternFill(start_color= "351c75", end_color= "351c75", fill_type="solid")   
                    cell_media_role_projection.font = Font(color="FFFFFF")  # White font color    

                if cell_media_role_stream.value == "STREAM":
                    cell_media_role_stream.fill = PatternFill(start_color= "d9d2e9", end_color= "d9d2e9", fill_type="solid")   
                    cell_media_role_stream.font = Font(color="000000")  # White font color    

            # Auto-adjust column widths to fit the content
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name

                #Skip column 4 when adjusting widths
                if column == 'D':
                    continue

                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2  # Adjust the width slightly to avoid cutting off text
                sheet.column_dimensions[column].width = adjusted_width

            for row in range(3, len(team_members) +3):
                member_cell = sheet.cell(row=row, column=1)
                member_cell.fill = PatternFill (start_color= "ff9900", end_color= "ff9900", fill_type="solid")

            # Add wrapping specifically for note rows if necessary
            note_start_row = len(data) - 4
            sheet.column_dimensions['D'].width = 20
            for note_row in range (note_start_row, len(data) - 1):
                
                cell = sheet.cell(row=note_row + 1, column=4)
                cell.alignment = Alignment(wrap_text=True)



                if cell.value:
                  if 'CAMERA' in cell.value.upper():
                    cell.fill = PatternFill(start_color="9933ff", end_color="9933ff", fill_type="solid")
                    cell.font = Font(color="FFFFFF")  # White font color
                  elif 'PROJECTION' in cell.value.upper():
                    cell.fill = PatternFill(start_color="351c75", end_color="351c75", fill_type="solid")
                    cell.font = Font(color="FFFFFF")  # White font color
                  elif 'STREAM' in cell.value.upper():
                    cell.fill = PatternFill(start_color="d9d2e9", end_color="d9d2e9", fill_type="solid")
                    cell.font = Font(color="000000")  # Black font color


                for col in range(5, len(sundays)* 3 + 2):
                    sheet.cell(row=note_row, column=col).alignment = Alignment(wrap_text=True)    




        # Save the workbook with merged cells, bold headers, text wrapping, and adjusted column widths
        workbook.save(output_path)

        print(f'All sheets saved to {output_path}')
    except PermissionError:
        print(f"Permission denied: Unable to write to {output_path}. Please ensure the file is closed and try again.")
    except Exception as e:
        print(f"An error occurred: {e}")


def load_team_members(file_name='team_members.csv'):
    """Loads team members from a CSV file."""
    if os.path.exists(file_name):
        df = pd.read_csv(file_name)
        return df['Members'].tolist(), True  # Return list and success flag
    return [], False  # Return empty list and failure flag

def save_team_members(team_members, file_name='team_members.csv'):
    """Saves team members to a CSV file."""
    df = pd.DataFrame({'Members': team_members})
    df.to_csv(file_name, index=False)

def main():
    print("Welcome to the Media Team Duty Roster Creator!")

    team_members, found = load_team_members()

    if found:
        print("Loaded team members:")
        for member in team_members:
            print(f" - {member}")
    else:
        print("No previous members found. Please enter the names of the media team members.")

    while True:
        member = input("Please enter the name of a media team member (or type 'done' when finished): ").strip()
        if member.lower() == 'done':
            break
        if member and member not in team_members:  # Avoid adding duplicates
            team_members.append(member)

    if not team_members:
        print("No team members provided. Exiting.")
        return

    # Save the team members to a CSV file
    save_team_members(team_members)

    # User inputs for year
    while True:
        year_input = input("Please enter the year (e.g., 2024): ")
        if year_input.isdigit() and len(year_input) == 4:
            year = int(year_input)
            break
        else:
            print("Invalid input. Please enter a valid year (4 digits).")

    # Get all Sundays and assign members
    roster = {}
    for month in range(1, 13):
        sundays = get_sundays(year, month)
        if sundays:
            monthly_roster = assign_members(sundays, team_members)
            roster.update(monthly_roster)

    create_excel_roster(year, roster, team_members)

if __name__ == "__main__":
    main()
